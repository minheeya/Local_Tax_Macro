# 지방세 신고 대상 엑셀 파일을 읽어서 순서대로 실행

import openpyxl
import singo
import re


# 필요한 초기 정보 입력
file_path = input("===== 대상자 파일 경로 입력:") + ".xlsx"
start_num = int(input("===== 시작 행 번호 입력: ").replace(" ", ""))
end_num = int(input("===== 마지막 행 번호 입력: ").replace(" ", ""))


# 엑셀 항목의 첫 번째 시트에 접근
try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
except Exception as e:
    print(e)
    print("===== 파일 제목과 시트를 확인해주세요 =====\n")


# 카테고리는 1번째 row에 존재
cate_tp = ws[1]


# 각 카테고리의 위치 저장
cate_idx_dic = {}
for idx in range(0, len(cate_tp)):
    cate_idx_dic[cate_tp[idx].value.replace(" ", "")] = idx


# 가상계좌 안나온 대상 목록
vtl_acct_wait_list = []


# 엑셀에 있는 정보로 지방세 신고 실행
row_num = start_num
while (row_num <= end_num):

    print("\n=====", end_num - row_num + 1, "개 남았습니다! ", "시작합니다 =====\n")

    target_info_tp = ws[row_num]
    target_info_dic = {
         
         "USER_NM":         target_info_tp[cate_idx_dic.get("이름")].value,
         "HP_NO":           target_info_tp[cate_idx_dic.get("연락처")].value.replace("-", ""),
         "CORP_NM":         target_info_tp[cate_idx_dic.get("사업장명")].value,
         "CORP_NO":         target_info_tp[cate_idx_dic.get("사업자번호")].value,
         "CORP_ADDR":       target_info_tp[cate_idx_dic.get("도로명주소")].value,
         "CORP_LB_ADDR":    target_info_tp[cate_idx_dic.get("지번주소")].value,
         "HTX_REG_NO":      target_info_tp[cate_idx_dic.get("홈택스접수번호")].value,
         "TOT_HTAX_AMT":    target_info_tp[cate_idx_dic.get("원천세신고금액")].value,
         "UUID":            target_info_tp[0].value.replace(" ", "")

    }

    target = singo.CorpSingo(target_info_dic)
    target.singo()


    print("===== 제가 할 수 있는 것은 다 완료 했어요 :) =====\n===== 신고 정보를 확인하고 저장해 주세요! =====\n")
    print(row_num, " row 사업장")
    print("이름: "              , target.user_nm)
    print("연락처: "            , target.hp_no)
    print("사업장명: "          , target.corp_nm)
    print("사업자번호: "        , target.corp_no)
    print("도로명주소: "        , target.corp_addr)
    print("도로명주소: "        , target.corp_lb_addr)
    print("홈택스접수번호: "    , target.htx_reg_no)
    print("원천세신고금액: "    , target.tot_htax_amt)
    print("\n=====================\n")

    
    vtl_acct_yn = int(re.sub(r'[^0-9]', '', input("1. 가상계좌 조회 완료\n2. 가상계좌 조회 미완료\n3. 신고 실패\n입력: ")))
    if vtl_acct_yn == 2:
        vtl_acct_wait_list.append({"이름": target_info_dic.get("USER_NM"), "UUID": target_info_dic.get("UUID")})
    
    next_step_num = int(re.sub(r'[^0-9]', '', input("\n1. 다음 신고 시작!\n2. 재시도 (망할신고연계오류)\n3. 종료\n입력: ")))

    if next_step_num == 3:
        break

    row_num = row_num if next_step_num == 2 else row_num + 1


print("\n===== 모두 끝났습니다! =====\n")

print("\n===== 가상계좌 누락 대상자 목록 =====\n")
for vtl_acct_wait in vtl_acct_wait_list:
    print(vtl_acct_wait, "\n")

print("\n===== 껀비 화이팅 😊 =====")




    


  

